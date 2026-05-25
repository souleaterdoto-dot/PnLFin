"""Excel import/export helpers for referral rate conditions."""

from __future__ import annotations

from dataclasses import dataclass, field
from datetime import date, datetime
from pathlib import Path
from typing import Any

from openpyxl import Workbook, load_workbook
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.styles import Font, PatternFill
from openpyxl.worksheet.worksheet import Worksheet

from app.domain.rate_models import RateCondition
from app.repositories.referrals_repository import ReferralsRepository
from app.services.rate_conditions_service import RateConditionsService
from app.services.rate_conditions_validation_service import RateConditionOverlapError


AMOUNT_BASIS_EQ = "usd_equivalent"
AMOUNT_BASIS_DEAL = "deal_currency"

HEADERS = (
    "\u0420\u0435\u0444\u0435\u0440\u0430\u043b",
    "\u0422\u0438\u043f \u0441\u0434\u0435\u043b\u043a\u0438",
    "\u0412\u0430\u043b\u044e\u0442\u0430",
    "\u041e\u0442",
    "\u0414\u043e",
    "\u041f\u0440\u043e\u0446\u0435\u043d\u0442, %",
    "\u0424\u0438\u043a\u0441\u0438\u0440\u043e\u0432\u0430\u043d\u043d\u0430\u044f \u0441\u0443\u043c\u043c\u0430",
    "\u0414\u0430\u0442\u0430 \u043d\u0430\u0447\u0430\u043b\u0430",
    "\u0414\u0430\u0442\u0430 \u043e\u043a\u043e\u043d\u0447\u0430\u043d\u0438\u044f",
)


@dataclass(slots=True)
class RateConditionsImportResult:
    """Result of importing rate conditions from Excel."""

    source_file: str
    amount_basis: str
    rows_total: int = 0
    rows_success: int = 0
    rows_failed: int = 0
    errors: list[str] = field(default_factory=list)


@dataclass(frozen=True, slots=True)
class _PreparedCondition:
    """Parsed condition waiting for all-file validation before saving."""

    row_number: int
    referral_name: str
    condition: RateCondition


class RateConditionsExcelService:
    """Import rate conditions from a constrained Excel template."""

    def __init__(
        self,
        referrals_repository: ReferralsRepository,
        rate_conditions_service: RateConditionsService,
    ) -> None:
        self._referrals_repository = referrals_repository
        self._rate_conditions_service = rate_conditions_service

    def import_file(self, file_path: str | Path) -> RateConditionsImportResult:
        """Import rate conditions from a read-only Excel file."""
        source_path = Path(file_path)
        workbook = load_workbook(source_path, read_only=True, data_only=True)
        try:
            sheet = workbook.active
            amount_basis = _parse_amount_basis(sheet["A1"].value)
            header_row = _find_header_row(sheet)
            column_map = _header_map(sheet, header_row)
            _validate_headers(column_map)
            result = RateConditionsImportResult(source_file=str(source_path), amount_basis=amount_basis)
            prepared: list[_PreparedCondition] = []
            for row_number in range(header_row + 1, sheet.max_row + 1):
                payload = {
                    header: sheet.cell(row=row_number, column=column).value
                    for header, column in column_map.items()
                }
                if _is_empty_row(payload):
                    continue
                result.rows_total += 1
                try:
                    prepared.append(self._prepare_row(payload, row_number, amount_basis, source_path.name))
                except Exception as exc:
                    result.rows_failed += 1
                    result.errors.append(f"\u0421\u0442\u0440\u043e\u043a\u0430 {row_number}: {exc}")
            result.errors.extend(_find_internal_conflicts(prepared))
            if result.errors:
                result.rows_failed = max(result.rows_failed, result.rows_total - result.rows_success)
                return result
            for item in prepared:
                try:
                    self._rate_conditions_service.save(item.condition)
                    result.rows_success += 1
                except RateConditionOverlapError as exc:
                    result.rows_failed += 1
                    result.errors.append(f"\u0421\u0442\u0440\u043e\u043a\u0430 {item.row_number}: {exc}")
                except Exception as exc:
                    result.rows_failed += 1
                    result.errors.append(f"\u0421\u0442\u0440\u043e\u043a\u0430 {item.row_number}: {exc}")
            return result
        finally:
            workbook.close()

    def create_example_file(self, target_path: str | Path) -> Path:
        """Create an Excel example file and return its path."""
        path = Path(target_path)
        path.parent.mkdir(parents=True, exist_ok=True)
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Rates"
        sheet["A1"] = "\u0412 \u0432\u0430\u043b\u044e\u0442\u0435 \u0441\u0434\u0435\u043b\u043a\u0438"
        sheet["A1"].font = Font(bold=True, color="FFFFFF")
        sheet["A1"].fill = PatternFill("solid", fgColor="2563EB")
        validation = DataValidation(
            type="list",
            formula1='"eq.,\u0412 \u0432\u0430\u043b\u044e\u0442\u0435 \u0441\u0434\u0435\u043b\u043a\u0438"',
            allow_blank=False,
        )
        validation.error = "\u0412\u044b\u0431\u0435\u0440\u0438\u0442\u0435 eq. \u0438\u043b\u0438 \u0412 \u0432\u0430\u043b\u044e\u0442\u0435 \u0441\u0434\u0435\u043b\u043a\u0438"
        validation.prompt = "\u0420\u0435\u0436\u0438\u043c \u0434\u0438\u0430\u043f\u0430\u0437\u043e\u043d\u0430 \u0441\u0443\u043c\u043c\u044b"
        sheet.add_data_validation(validation)
        validation.add(sheet["A1"])
        for index, header in enumerate(HEADERS, start=1):
            cell = sheet.cell(row=2, column=index, value=header)
            cell.font = Font(bold=True, color="0F172A")
            cell.fill = PatternFill("solid", fgColor="DBEAFE")
        values = [
            "\u0410\u043b\u044c\u0444\u0430 \u0431\u0430\u043d\u043a",
            "\u0418\u043c\u043f\u043e\u0440\u0442",
            "USD",
            "-",
            10000,
            0,
            300,
            date(2026, 5, 1),
            date(2030, 5, 1),
        ]
        for index, value in enumerate(values, start=1):
            sheet.cell(row=3, column=index, value=value)
        sheet["D3"].number_format = '#,##0.00'
        sheet["E3"].number_format = '#,##0.00'
        sheet["F3"].number_format = '0.00%'
        sheet["G3"].number_format = '#,##0.00'
        sheet["H3"].number_format = "DD.MM.YYYY"
        sheet["I3"].number_format = "DD.MM.YYYY"
        widths = [22, 16, 12, 14, 14, 14, 24, 16, 18]
        for index, width in enumerate(widths, start=1):
            sheet.column_dimensions[sheet.cell(row=2, column=index).column_letter].width = width
        workbook.save(path)
        workbook.close()
        return path

    def _prepare_row(
        self,
        payload: dict[str, Any],
        row_number: int,
        amount_basis: str,
        source_file_name: str,
    ) -> _PreparedCondition:
        referral_name = _required_text(payload[HEADERS[0]], "\u0420\u0435\u0444\u0435\u0440\u0430\u043b")
        referral = self._referrals_repository.find_by_name_or_code(referral_name)
        if referral is None or referral.id is None:
            raise ValueError(f"\u043d\u0435\u043a\u043e\u0440\u0440\u0435\u043a\u0442\u043d\u044b\u0439 \u0440\u0435\u0444\u0435\u0440\u0430\u043b: '{referral_name}'")

        operation_type = _operation_type(payload[HEADERS[1]])
        currency = _currency(payload[HEADERS[2]])
        amount_from = _optional_float(payload[HEADERS[3]], HEADERS[3])
        amount_to = _optional_float(payload[HEADERS[4]], HEADERS[4])
        if amount_from is not None and amount_from < 0:
            raise ValueError("\u041e\u0442 \u043d\u0435 \u043c\u043e\u0436\u0435\u0442 \u0431\u044b\u0442\u044c \u043c\u0435\u043d\u044c\u0448\u0435 0")
        if amount_to is not None and amount_to <= 0:
            raise ValueError("\u0414\u043e \u0434\u043e\u043b\u0436\u043d\u043e \u0431\u044b\u0442\u044c \u0431\u043e\u043b\u044c\u0448\u0435 0")
        if amount_from is not None and amount_to is not None and amount_from >= amount_to:
            raise ValueError("\u041e\u0442 \u0434\u043e\u043b\u0436\u043d\u043e \u0431\u044b\u0442\u044c \u043c\u0435\u043d\u044c\u0448\u0435 \u0414\u043e")
        date_from = _optional_date(payload[HEADERS[7]], HEADERS[7])
        date_to = _optional_date(payload[HEADERS[8]], HEADERS[8])
        if date_from and date_to and date_from > date_to:
            raise ValueError("\u0414\u0430\u0442\u0430 \u043d\u0430\u0447\u0430\u043b\u0430 \u043f\u043e\u0437\u0436\u0435 \u0414\u0430\u0442\u044b \u043e\u043a\u043e\u043d\u0447\u0430\u043d\u0438\u044f")
        rate_value = _percent(payload[HEADERS[5]], HEADERS[5])
        fixed_amount = _optional_float(payload[HEADERS[6]], HEADERS[6])
        if fixed_amount is not None and fixed_amount < 0:
            raise ValueError("\u0424\u0438\u043a\u0441\u0438\u0440\u043e\u0432\u0430\u043d\u043d\u0430\u044f \u0441\u0443\u043c\u043c\u0430 \u043d\u0435 \u043c\u043e\u0436\u0435\u0442 \u0431\u044b\u0442\u044c \u043c\u0435\u043d\u044c\u0448\u0435 0")
        condition = RateCondition(
            referral_id=int(referral.id),
            priority=row_number,
            operation_type=operation_type,
            currency=currency,
            amount_from=amount_from,
            amount_to=amount_to,
            amount_basis=amount_basis,
            date_from=date_from,
            date_to=date_to,
            rate_value=rate_value,
            percent_commission_currency=currency,
            fixed_commission_amount=fixed_amount,
            fixed_commission_currency=currency,
            commission_type="mixed",
            comment=f"{source_file_name}, строка {row_number}",
        )
        return _PreparedCondition(row_number=row_number, referral_name=referral_name, condition=condition)


def _parse_amount_basis(value: Any) -> str:
    text = str(value or "").strip().casefold()
    if text == "eq.":
        return AMOUNT_BASIS_EQ
    if text == "\u0432 \u0432\u0430\u043b\u044e\u0442\u0435 \u0441\u0434\u0435\u043b\u043a\u0438":
        return AMOUNT_BASIS_DEAL
    raise ValueError("A1 \u0434\u043e\u043b\u0436\u043d\u0430 \u0431\u044b\u0442\u044c 'eq.' \u0438\u043b\u0438 '\u0432 \u0432\u0430\u043b\u044e\u0442\u0435 \u0441\u0434\u0435\u043b\u043a\u0438'")


def _find_header_row(sheet: Worksheet) -> int:
    for row_number in range(1, min(sheet.max_row, 10) + 1):
        values = {str(sheet.cell(row=row_number, column=column).value or "").strip() for column in range(1, sheet.max_column + 1)}
        if HEADERS[0] in values and HEADERS[1] in values:
            return row_number
    raise ValueError("\u043d\u0435 \u043d\u0430\u0439\u0434\u0435\u043d\u0430 \u0441\u0442\u0440\u043e\u043a\u0430 \u0441 \u0437\u0430\u0433\u043e\u043b\u043e\u0432\u043a\u0430\u043c\u0438")


def _header_map(sheet: Worksheet, header_row: int) -> dict[str, int]:
    return {
        str(sheet.cell(row=header_row, column=column).value or "").strip(): column
        for column in range(1, sheet.max_column + 1)
    }


def _validate_headers(column_map: dict[str, int]) -> None:
    missing = [header for header in HEADERS if header not in column_map]
    if missing:
        raise ValueError("\u043d\u0435\u0442 \u043a\u043e\u043b\u043e\u043d\u043e\u043a: " + ", ".join(missing))


def _is_empty_row(payload: dict[str, Any]) -> bool:
    return all(str(value or "").strip() == "" for value in payload.values())


def _required_text(value: Any, field_name: str) -> str:
    text = str(value or "").strip()
    if not text or text == "-":
        raise ValueError(f"\u043f\u043e\u043b\u0435 '{field_name}' \u043f\u0443\u0441\u0442\u043e\u0435")
    return text


def _currency(value: Any) -> str:
    text = _required_text(value, "\u0412\u0430\u043b\u044e\u0442\u0430").upper()
    if not text.isalpha() or len(text) < 3 or len(text) > 5:
        raise ValueError("\u0412\u0430\u043b\u044e\u0442\u0430 \u0434\u043e\u043b\u0436\u043d\u0430 \u0431\u044b\u0442\u044c \u0442\u0435\u043a\u0441\u0442\u043e\u043c: USD, EUR, CNY, USDT")
    return text


def _operation_type(value: Any) -> str | None:
    text = str(value or "").strip().casefold()
    if not text:
        return None
    if text in {"\u0438\u043c\u043f\u043e\u0440\u0442", "import"}:
        return "import"
    if text in {"\u044d\u043a\u0441\u043f\u043e\u0440\u0442", "export"}:
        return "export"
    raise ValueError("\u0442\u0438\u043f \u0441\u0434\u0435\u043b\u043a\u0438 \u0434\u043e\u043b\u0436\u0435\u043d \u0431\u044b\u0442\u044c '\u0418\u043c\u043f\u043e\u0440\u0442' \u0438\u043b\u0438 '\u042d\u043a\u0441\u043f\u043e\u0440\u0442'")


def _optional_float(value: Any, field_name: str = "\u0447\u0438\u0441\u043b\u043e") -> float | None:
    if value is None:
        return None
    if isinstance(value, int | float):
        return float(value)
    text = str(value).strip().replace("\u00a0", " ").replace(" ", "")
    if text in {"", "-"}:
        return None
    try:
        return float(text.replace(",", "."))
    except ValueError as exc:
        raise ValueError(f"\u043f\u043e\u043b\u0435 '{field_name}' \u0434\u043e\u043b\u0436\u043d\u043e \u0431\u044b\u0442\u044c \u0447\u0438\u0441\u043b\u043e\u043c") from exc


def _percent(value: Any, field_name: str = "\u041f\u0440\u043e\u0446\u0435\u043d\u0442, %") -> float:
    if value is None:
        return 0.0
    if isinstance(value, int | float):
        number = float(value)
        return number * 100 if 0 < abs(number) <= 1 else number
    text = str(value).strip().replace("\u00a0", " ").replace(" ", "")
    if not text or text == "-":
        return 0.0
    has_percent = "%" in text
    try:
        number = float(text.replace("%", "").replace(",", "."))
    except ValueError as exc:
        raise ValueError(f"\u043f\u043e\u043b\u0435 '{field_name}' \u0434\u043e\u043b\u0436\u043d\u043e \u0431\u044b\u0442\u044c \u0447\u0438\u0441\u043b\u043e\u043c \u0438\u043b\u0438 %") from exc
    return number if has_percent else number


def _optional_date(value: Any, field_name: str = "\u0434\u0430\u0442\u0430") -> str | None:
    if value is None:
        return None
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()
    text = str(value).strip()
    if not text or text == "-":
        return None
    for fmt in ("%d.%m.%Y", "%Y-%m-%d", "%d/%m/%Y"):
        try:
            return datetime.strptime(text, fmt).date().isoformat()
        except ValueError:
            continue
    try:
        return date.fromisoformat(text).isoformat()
    except ValueError as exc:
        raise ValueError(f"\u043f\u043e\u043b\u0435 '{field_name}' \u0434\u043e\u043b\u0436\u043d\u043e \u0431\u044b\u0442\u044c \u0434\u0430\u0442\u043e\u0439") from exc


def _find_internal_conflicts(items: list[_PreparedCondition]) -> list[str]:
    errors: list[str] = []
    for left_index, left in enumerate(items):
        for right in items[left_index + 1:]:
            if left.condition.referral_id != right.condition.referral_id:
                continue
            if _conditions_overlap(left.condition, right.condition):
                errors.append(
                    f"Строка {right.row_number}: конфликт условий ставок. "
                    f"Она может подойти под ту же сделку, что и строка {left.row_number}. "
                    f"Строка {left.row_number}: {_condition_summary(left.condition)}. "
                    "Измените диапазон суммы, период дат, валюту или тип сделки."
                )
    return errors


def _conditions_overlap(left: RateCondition, right: RateCondition) -> bool:
    return (
        _amount_basis(left) == _amount_basis(right)
        and _wildcard_text_overlap(left.operation_type, right.operation_type)
        and _wildcard_text_overlap(left.currency, right.currency)
        and _range_overlap(left.amount_from, left.amount_to, right.amount_from, right.amount_to)
        and _date_range_overlap(left.date_from, left.date_to, right.date_from, right.date_to)
    )


def _wildcard_text_overlap(left: str | None, right: str | None) -> bool:
    if str(left or "").strip() in {"", "0"} or str(right or "").strip() in {"", "0"}:
        return True
    return str(left).strip().casefold() == str(right).strip().casefold()


def _range_overlap(
    left_from: float | None,
    left_to: float | None,
    right_from: float | None,
    right_to: float | None,
) -> bool:
    start_left = float(left_from or 0)
    end_left = float("inf") if left_to is None else float(left_to)
    start_right = float(right_from or 0)
    end_right = float("inf") if right_to is None else float(right_to)
    return start_left < end_right and start_right < end_left


def _date_range_overlap(
    left_from: str | None,
    left_to: str | None,
    right_from: str | None,
    right_to: str | None,
) -> bool:
    start_left = _date_or_default(left_from, date.min)
    end_left = _date_or_default(left_to, date.max)
    start_right = _date_or_default(right_from, date.min)
    end_right = _date_or_default(right_to, date.max)
    return start_left <= end_right and start_right <= end_left


def _date_or_default(value: str | None, default: date) -> date:
    if not value:
        return default
    return date.fromisoformat(value)


def _condition_summary(condition: RateCondition) -> str:
    operation_type = _operation_type_label(condition.operation_type)
    currency = condition.currency or "любая валюта"
    amount_from = _fmt_amount(condition.amount_from or 0)
    amount_to = _fmt_amount(condition.amount_to) if condition.amount_to is not None else "без лимита"
    amount_basis = "eq. USD" if _amount_basis(condition) == AMOUNT_BASIS_EQ else "в валюте сделки"
    date_from = _fmt_date(condition.date_from) if condition.date_from else "с начала"
    date_to = _fmt_date(condition.date_to) if condition.date_to else "бессрочно"
    return f"{operation_type}, {currency}, сумма {amount_from}-{amount_to} ({amount_basis}), период {date_from}-{date_to}"


def _operation_type_label(value: str | None) -> str:
    normalized = str(value or "").strip().casefold()
    if normalized == "import":
        return "Импорт"
    if normalized == "export":
        return "Экспорт"
    return "любой тип"


def _fmt_amount(value: float | None) -> str:
    if value is None:
        return "без лимита"
    return f"{float(value):,.0f}".replace(",", " ")


def _fmt_date(value: str | None) -> str:
    if not value:
        return ""
    try:
        return date.fromisoformat(value).strftime("%d.%m.%Y")
    except ValueError:
        return value


def _amount_basis(condition: RateCondition) -> str:
    return condition.amount_basis if condition.amount_basis in {AMOUNT_BASIS_DEAL, AMOUNT_BASIS_EQ} else AMOUNT_BASIS_DEAL
