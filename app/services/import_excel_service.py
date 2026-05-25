"""Power Query-safe Excel import service.

The service treats Excel files as read-only data sources. It never saves the
workbook, never refreshes Power Query, and only reads already materialized data
from Excel Table "All" or worksheet "All".
"""

from __future__ import annotations

import json
import posixpath
import zipfile
from datetime import date, datetime
from pathlib import Path
from typing import Any
from xml.etree import ElementTree

from openpyxl import load_workbook
from openpyxl.utils.cell import range_boundaries

from app.domain.import_models import (
    ExcelRow,
    ExcelSource,
    ImportBatch,
    ImportErrorRecord,
    ImportResult,
)
from app.domain.models import Deal, ValidationError
from app.repositories.deals_repository import DealsRepository
from app.repositories.import_batches_repository import ImportBatchesRepository


COLUMN_MAPPING = {
    "№ сделки": "external_deal_id",
    "Менеджер": "manager",
    "Повторный платёж (чекбокс)": "is_repeat_payment",
    "Повторный платеж (чекбокс)": "is_repeat_payment",
    "Дата поступления заявки": "request_date",
    "Дата фиксации с клиентом": "client_fix_date",
    "Дата списания с баланса ПА": "agent_writeoff_date",
    "Дата получения клиентом": "client_receive_date",
    "Возврат (чекбокс)": "is_refund",
    "Дата возврата средств на ПА": "agent_refund_date",
    "Дата возврата средств клиенту": "client_refund_date",
    "Статус платежа": "payment_status",
    "Название клиента": "client_name",
    "Компания получатель": "receiver_company",
    "Страна банка получателя": "receiver_bank_country",
    "Сумма сделки": "deal_amount",
    "Валюта сделки": "deal_currency",
    "Ставка для клиента (%)": "client_rate_percent",
    "Фикс. Комиссия (сумма)": "fixed_commission_amount",
    "Фикс. Комиссия (валюта)": "fixed_commission_currency",
    "SWIFT (сумма)": "swift_amount",
    "SWIFT (валюта)": "swift_currency",
    "Курс фиксации с клиентом": "client_fix_rate",
    "Курс к USD": "usd_rate",
    "Кросс-курс с клиентом": "client_cross_rate",
    "Платежный агент": "payment_agent",
    "Комиссия ПА (сумма)": "agent_commission_amount",
    "Валюта комиссии ПА": "agent_commission_currency",
    "Комиссия за свифт ПА (сумма)": "swift_commission_amount",
    "Валюта комиссии за свифт ПА": "swift_commission_currency",
    "Статья название (клиент)": "customer_article_name",
    "trade_date": "trade_date",
    "value_date": "value_date",
    "operation_type": "operation_type",
    "counterparty": "counterparty",
    "currency_buy": "currency_buy",
    "amount_buy": "amount_buy",
    "currency_sell": "currency_sell",
    "amount_sell": "amount_sell",
    "rate_fact": "rate_fact",
    "commission": "commission",
    "portfolio": "portfolio",
    "comment": "comment",
    "included_in_calc": "included_in_calc",
}


NATIVE_REQUIRED_COLUMNS = {
    "trade_date",
    "value_date",
    "operation_type",
    "counterparty",
    "currency_buy",
    "amount_buy",
    "currency_sell",
    "amount_sell",
    "rate_fact",
    "portfolio",
}
POWER_QUERY_REQUIRED_COLUMNS = {"client_name", "deal_amount", "deal_currency"}
POWER_QUERY_DATE_COLUMNS = {
    "client_fix_date",
    "request_date",
    "agent_writeoff_date",
    "client_receive_date",
}
POWER_QUERY_RATE_COLUMNS = {"client_fix_rate", "client_cross_rate", "usd_rate"}
SUPPORTED_EXTENDED_CURRENCIES = {"USDT"}
POWER_QUERY_AMOUNT_COLUMNS = {
    "deal_amount",
    "fixed_commission_amount",
    "swift_amount",
    "agent_commission_amount",
    "swift_commission_amount",
}


class ImportSourceSelectionRequired(RuntimeError):
    """Raised when the workbook has no All table/sheet and UI must ask user."""

    def __init__(self, file_path: str, available_sheets: list[str]) -> None:
        super().__init__("Table or sheet 'All' was not found. Select a sheet to import.")
        self.file_path = file_path
        self.available_sheets = available_sheets


class ImportExcelService:
    """Read a prepared Excel result table and replace the deals registry."""

    def __init__(
        self,
        deals_repository: DealsRepository | None = None,
        batches_repository: ImportBatchesRepository | None = None,
    ) -> None:
        self._deals_repository = deals_repository or DealsRepository()
        self._batches_repository = batches_repository or ImportBatchesRepository()

    def import_file(self, file_path: str, selected_sheet: str | None = None) -> ImportResult:
        """Import Excel Table "All", worksheet "All", or a user-selected sheet."""
        source = self.resolve_source(file_path, selected_sheet)
        batch_id = self._batches_repository.create(
            ImportBatch(
                source_file=source.source_file,
                source_sheet=source.source_sheet,
                status="running",
            )
        )

        try:
            headers, rows = self._read_rows(source)
            column_errors = self._validate_columns(headers)
            if column_errors:
                self._batches_repository.add_errors(
                    self._to_error_records(batch_id, column_errors, {})
                )
                message = "; ".join(error.message for error in column_errors)
                self._batches_repository.update_status(
                    batch_id=batch_id,
                    rows_total=0,
                    rows_success=0,
                    rows_failed=0,
                    status="failed",
                    error_message=message,
                )
                return ImportResult(
                    import_batch_id=batch_id,
                    source_file=source.source_file,
                    source_sheet=source.source_sheet,
                    rows_total=0,
                    rows_success=0,
                    rows_failed=0,
                    status="failed",
                    errors=column_errors,
                    error_message=message,
                )

            deals: list[Deal] = []
            validation_errors: list[ValidationError] = []
            raw_payloads_by_row: dict[int, dict[str, Any]] = {}
            for row in rows:
                raw_payloads_by_row[row.source_row_number] = row.raw_payload
                deal, errors = self._row_to_deal(row, source, batch_id)
                if errors:
                    validation_errors.extend(errors)
                    continue
                if deal:
                    deals.append(deal)

            self._deals_repository.delete_all()
            imported_count = self._deals_repository.add_many(deals) if deals else 0
            failed_rows = len({error.row_number for error in validation_errors})
            if validation_errors:
                error_records: list[ImportErrorRecord] = []
                for error in validation_errors:
                    payload = raw_payloads_by_row.get(error.row_number, {})
                    error_records.extend(self._to_error_records(batch_id, [error], payload))
                self._batches_repository.add_errors(error_records)

            status = "success" if not validation_errors else "partial"
            self._batches_repository.update_status(
                batch_id=batch_id,
                rows_total=len(rows),
                rows_success=imported_count,
                rows_failed=failed_rows,
                status=status,
            )
            return ImportResult(
                import_batch_id=batch_id,
                source_file=source.source_file,
                source_sheet=source.source_sheet,
                rows_total=len(rows),
                rows_success=imported_count,
                rows_failed=failed_rows,
                status=status,
                errors=validation_errors,
            )
        except Exception as exc:
            self._batches_repository.update_status(
                batch_id=batch_id,
                rows_total=0,
                rows_success=0,
                rows_failed=0,
                status="failed",
                error_message=str(exc),
            )
            raise

    def resolve_source(self, file_path: str, selected_sheet: str | None = None) -> ExcelSource:
        """Resolve import source without modifying workbook connections or queries."""
        path = Path(file_path)
        if path.suffix.lower() not in {".xlsx", ".xlsm"}:
            raise ValueError("Only .xlsx and .xlsm files are supported for safe read-only import.")

        available_sheets = self._list_sheets(path)
        table_source = self._find_table_all(path, available_sheets)
        if table_source:
            return table_source

        if selected_sheet:
            if selected_sheet not in available_sheets:
                raise ValueError(f"Sheet '{selected_sheet}' was not found in workbook.")
            return ExcelSource(
                source_file=str(path),
                source_sheet=selected_sheet,
                source_type="sheet",
                available_sheets=tuple(available_sheets),
            )

        for sheet_name in available_sheets:
            if sheet_name.lower() == "all":
                return ExcelSource(
                    source_file=str(path),
                    source_sheet=sheet_name,
                    source_type="sheet",
                    available_sheets=tuple(available_sheets),
                )

        raise ImportSourceSelectionRequired(str(path), available_sheets)

    def _read_rows_from_workbook(self, file_path: str, source: ExcelSource) -> tuple[list[str], list[ExcelRow]]:
        workbook = load_workbook(
            filename=file_path,
            read_only=True,
            data_only=True,
            keep_links=True,
        )
        try:
            worksheet = workbook[source.source_sheet]
            if source.cell_range:
                min_col, min_row, max_col, max_row = range_boundaries(source.cell_range)
                row_iter = worksheet.iter_rows(
                    min_row=min_row,
                    max_row=max_row,
                    min_col=min_col,
                    max_col=max_col,
                    values_only=True,
                )
                header_row_number = min_row
            else:
                row_iter = worksheet.iter_rows(values_only=True)
                header_row_number = 1

            try:
                header_values = next(row_iter)
            except StopIteration:
                return [], []

            headers = [str(value).strip() if value is not None else "" for value in header_values]
            rows: list[ExcelRow] = []
            for offset, values in enumerate(row_iter, start=header_row_number + 1):
                raw_payload = {
                    header: value for header, value in zip(headers, values, strict=False) if header
                }
                if self._is_empty_row(raw_payload):
                    continue
                normalized = {
                    self._normalize_header(header): value
                    for header, value in raw_payload.items()
                    if self._normalize_header(header)
                }
                rows.append(
                    ExcelRow(
                        source_row_number=offset,
                        raw_payload=raw_payload,
                        normalized_payload=normalized,
                    )
                )
            return headers, rows
        finally:
            workbook.close()

    def _validate_columns(self, headers: list[str]) -> list[ValidationError]:
        normalized_headers = {self._normalize_header(header) for header in headers if header}
        missing_native = sorted(NATIVE_REQUIRED_COLUMNS - normalized_headers)
        has_native_schema = not missing_native

        missing_power_query = sorted(POWER_QUERY_REQUIRED_COLUMNS - normalized_headers)
        has_date = bool(normalized_headers & POWER_QUERY_DATE_COLUMNS)
        has_rate = bool(normalized_headers & POWER_QUERY_RATE_COLUMNS)
        has_power_query_schema = not missing_power_query and has_date and has_rate

        if has_native_schema or has_power_query_schema:
            return []

        errors = [
            ValidationError(
                1,
                "columns",
                "Required columns are missing. Provide native deal columns or Power Query columns.",
            )
        ]
        if missing_native:
            errors.append(
                ValidationError(1, "native_schema", f"Missing native columns: {', '.join(missing_native)}")
            )
        if missing_power_query or not has_date or not has_rate:
            details = list(missing_power_query)
            if not has_date:
                details.append("one of date columns: client_fix_date/request_date/agent_writeoff_date/client_receive_date")
            if not has_rate:
                details.append("one of rate columns: client_fix_rate/client_cross_rate/usd_rate")
            errors.append(
                ValidationError(1, "power_query_schema", f"Missing Power Query columns: {', '.join(details)}")
            )
        return errors

    def _row_to_deal(
        self,
        row: ExcelRow,
        source: ExcelSource,
        batch_id: int,
    ) -> tuple[Deal | None, list[ValidationError]]:
        payload = row.normalized_payload
        errors: list[ValidationError] = []
        trade_date = self._parse_date(
            self._first(payload, "trade_date", "client_fix_date", "request_date"),
            row.source_row_number,
            "trade_date",
            errors,
        )
        value_date = self._parse_date(
            self._first(payload, "value_date", "agent_writeoff_date", "client_receive_date")
            or trade_date,
            row.source_row_number,
            "value_date",
            errors,
        )
        amount_buy = self._parse_float(
            self._first(payload, "amount_buy", "deal_amount"),
            row.source_row_number,
            "amount_buy",
            errors,
        )
        is_export_amount = amount_buy is not None and amount_buy < 0
        if amount_buy is not None:
            amount_buy = abs(amount_buy)
        rate_fact = self._parse_float(
            self._first(payload, "rate_fact", "client_fix_rate", "client_cross_rate", "usd_rate"),
            row.source_row_number,
            "rate_fact",
            errors,
        )
        explicit_amount_sell = self._first(payload, "amount_sell")
        amount_sell = self._parse_float(
            explicit_amount_sell,
            row.source_row_number,
            "amount_sell",
            errors,
            allow_empty=True,
        )
        if amount_sell is not None:
            amount_sell = abs(amount_sell)
        if (amount_sell is None or amount_sell == 0) and amount_buy and rate_fact:
            amount_sell = amount_buy * rate_fact

        commission = self._parse_float(
            self._first(payload, "commission", "fixed_commission_amount"),
            row.source_row_number,
            "commission",
            errors,
            allow_empty=True,
        )
        if commission is not None:
            commission = abs(commission)
        excel_fields = self._extract_excel_header_fields(payload, row.source_row_number, errors)
        self._normalize_amount_fields(excel_fields)
        currency_buy = str(self._first(payload, "currency_buy", "deal_currency") or "").strip().upper()
        currency_sell = str(self._first(payload, "currency_sell") or "RUB").strip().upper()
        counterparty = str(self._first(payload, "counterparty", "client_name", "receiver_company") or "").strip()
        portfolio = str(self._first(payload, "portfolio", "payment_agent") or "Без субагента").strip()
        has_deal_amount = not _is_empty(payload.get("deal_amount"))
        if has_deal_amount and amount_buy is not None:
            operation_type = "EXPORT" if is_export_amount else "IMPORT"
        else:
            operation_type = str(self._first(payload, "operation_type", "payment_status") or "PAYMENT").strip()

        required_text = {
            "counterparty": counterparty,
            "currency_buy": currency_buy,
            "currency_sell": currency_sell,
            "portfolio": portfolio,
            "operation_type": operation_type,
        }
        for field_name, value in required_text.items():
            if not value:
                errors.append(ValidationError(row.source_row_number, field_name, "Required value is missing"))

        for field_name, currency in {"currency_buy": currency_buy, "currency_sell": currency_sell}.items():
            if currency and not self._is_supported_currency_code(currency):
                errors.append(
                    ValidationError(
                        row.source_row_number,
                        field_name,
                        "Currency must be a 3-letter ISO code or supported crypto code",
                        currency,
                    )
                )

        if amount_buy is not None and amount_buy <= 0:
            errors.append(ValidationError(row.source_row_number, "amount_buy", "Amount must be positive", amount_buy))
        if amount_sell is not None and amount_sell <= 0:
            errors.append(ValidationError(row.source_row_number, "amount_sell", "Amount must be positive", amount_sell))
        if rate_fact is not None and rate_fact <= 0:
            errors.append(ValidationError(row.source_row_number, "rate_fact", "Rate must be positive", rate_fact))
        if errors:
            return None, errors

        raw_payload_json = _json_dump(row.raw_payload)
        return (
            Deal(
                trade_date=trade_date or "",
                value_date=value_date or "",
                operation_type=operation_type,
                counterparty=counterparty,
                currency_buy=currency_buy,
                amount_buy=amount_buy or 0,
                currency_sell=currency_sell,
                amount_sell=amount_sell or 0,
                rate_fact=rate_fact or 0,
                commission=commission or 0,
                portfolio=portfolio,
                comment=self._build_comment(payload),
                external_deal_id=excel_fields["external_deal_id"],
                manager=excel_fields["manager"],
                is_repeat_payment=excel_fields["is_repeat_payment"],
                request_date=excel_fields["request_date"],
                client_fix_date=excel_fields["client_fix_date"],
                agent_writeoff_date=excel_fields["agent_writeoff_date"],
                client_receive_date=excel_fields["client_receive_date"],
                is_refund=excel_fields["is_refund"],
                agent_refund_date=excel_fields["agent_refund_date"],
                client_refund_date=excel_fields["client_refund_date"],
                payment_status=excel_fields["payment_status"],
                client_name=excel_fields["client_name"],
                receiver_company=excel_fields["receiver_company"],
                receiver_bank_country=excel_fields["receiver_bank_country"],
                deal_amount=excel_fields["deal_amount"],
                deal_currency=excel_fields["deal_currency"],
                client_rate_percent=excel_fields["client_rate_percent"],
                fixed_commission_amount=excel_fields["fixed_commission_amount"],
                fixed_commission_currency=excel_fields["fixed_commission_currency"],
                swift_amount=excel_fields["swift_amount"],
                swift_currency=excel_fields["swift_currency"],
                client_fix_rate=excel_fields["client_fix_rate"],
                usd_rate=excel_fields["usd_rate"],
                client_cross_rate=excel_fields["client_cross_rate"],
                payment_agent=excel_fields["payment_agent"],
                agent_commission_amount=excel_fields["agent_commission_amount"],
                agent_commission_currency=excel_fields["agent_commission_currency"],
                swift_commission_amount=excel_fields["swift_commission_amount"],
                swift_commission_currency=excel_fields["swift_commission_currency"],
                customer_article_name=excel_fields["customer_article_name"],
                source_file=source.source_file,
                source_sheet=source.source_sheet,
                source_row_number=row.source_row_number,
                import_batch_id=batch_id,
                raw_payload_json=raw_payload_json,
                included_in_calc=self._parse_bool(self._first(payload, "included_in_calc", default=True)),
            ),
            [],
        )

    def _extract_excel_header_fields(
        self,
        payload: dict[str, Any],
        row_number: int,
        errors: list[ValidationError],
    ) -> dict[str, Any]:
        """Extract every known Excel header field into typed deal columns."""
        return {
            "external_deal_id": self._optional_text(payload.get("external_deal_id")),
            "manager": self._optional_text(payload.get("manager")),
            "is_repeat_payment": self._optional_bool(
                payload.get("is_repeat_payment"),
                row_number,
                "is_repeat_payment",
                errors,
            ),
            "request_date": self._parse_optional_date(payload.get("request_date"), row_number, "request_date", errors),
            "client_fix_date": self._parse_optional_date(payload.get("client_fix_date"), row_number, "client_fix_date", errors),
            "agent_writeoff_date": self._parse_optional_date(payload.get("agent_writeoff_date"), row_number, "agent_writeoff_date", errors),
            "client_receive_date": self._parse_optional_date(payload.get("client_receive_date"), row_number, "client_receive_date", errors),
            "is_refund": self._optional_bool(
                payload.get("is_refund"),
                row_number,
                "is_refund",
                errors,
            ),
            "agent_refund_date": self._parse_optional_date(payload.get("agent_refund_date"), row_number, "agent_refund_date", errors),
            "client_refund_date": self._parse_optional_date(payload.get("client_refund_date"), row_number, "client_refund_date", errors),
            "payment_status": self._optional_text(payload.get("payment_status")),
            "client_name": self._optional_text(payload.get("client_name")),
            "receiver_company": self._optional_text(payload.get("receiver_company")),
            "receiver_bank_country": self._optional_text(payload.get("receiver_bank_country")),
            "deal_amount": self._parse_optional_float(payload.get("deal_amount"), row_number, "deal_amount", errors),
            "deal_currency": self._optional_upper(payload.get("deal_currency")),
            "client_rate_percent": self._parse_optional_positive_float(
                payload.get("client_rate_percent"),
                row_number,
                "client_rate_percent",
                errors,
            ),
            "fixed_commission_amount": self._parse_optional_float(payload.get("fixed_commission_amount"), row_number, "fixed_commission_amount", errors),
            "fixed_commission_currency": self._optional_upper(payload.get("fixed_commission_currency")),
            "swift_amount": self._parse_optional_float(payload.get("swift_amount"), row_number, "swift_amount", errors),
            "swift_currency": self._optional_upper(payload.get("swift_currency")),
            "client_fix_rate": self._parse_optional_float(payload.get("client_fix_rate"), row_number, "client_fix_rate", errors),
            "usd_rate": self._parse_optional_float(payload.get("usd_rate"), row_number, "usd_rate", errors),
            "client_cross_rate": self._parse_optional_float(payload.get("client_cross_rate"), row_number, "client_cross_rate", errors),
            "payment_agent": self._optional_text(payload.get("payment_agent")),
            "agent_commission_amount": self._parse_optional_float(payload.get("agent_commission_amount"), row_number, "agent_commission_amount", errors),
            "agent_commission_currency": self._optional_upper(payload.get("agent_commission_currency")),
            "swift_commission_amount": self._parse_optional_float(payload.get("swift_commission_amount"), row_number, "swift_commission_amount", errors),
            "swift_commission_currency": self._optional_upper(payload.get("swift_commission_currency")),
            "customer_article_name": self._optional_text(payload.get("customer_article_name")),
        }

    def _normalize_amount_fields(self, fields: dict[str, Any]) -> None:
        """Store imported money amounts as positive values."""
        for field_name in POWER_QUERY_AMOUNT_COLUMNS:
            value = fields.get(field_name)
            if value is not None:
                fields[field_name] = abs(float(value))

    def _read_rows(self, source: ExcelSource) -> tuple[list[str], list[ExcelRow]]:
        source_path = self._source_path(source.source_file)
        return self._read_rows_from_workbook(str(source_path), source)

    def _source_path(self, source_file: str) -> Path:
        path = Path(source_file)
        if path.exists():
            return path
        return Path.cwd() / source_file

    def _list_sheets(self, path: Path) -> list[str]:
        workbook = load_workbook(
            filename=str(path),
            read_only=True,
            data_only=True,
            keep_links=True,
        )
        try:
            return list(workbook.sheetnames)
        finally:
            workbook.close()

    def _find_table_all(self, path: Path, available_sheets: list[str]) -> ExcelSource | None:
        with zipfile.ZipFile(path) as archive:
            try:
                workbook_xml = ElementTree.fromstring(archive.read("xl/workbook.xml"))
                rels_xml = ElementTree.fromstring(archive.read("xl/_rels/workbook.xml.rels"))
            except KeyError:
                return None

            rels = {
                rel.attrib["Id"]: _resolve_part("xl/workbook.xml", rel.attrib["Target"])
                for rel in rels_xml
            }
            for sheet in workbook_xml.findall(".//{http://schemas.openxmlformats.org/spreadsheetml/2006/main}sheet"):
                sheet_name = sheet.attrib.get("name")
                relation_id = sheet.attrib.get("{http://schemas.openxmlformats.org/officeDocument/2006/relationships}id")
                worksheet_part = rels.get(relation_id or "")
                if not sheet_name or not worksheet_part:
                    continue
                table = self._find_table_all_in_sheet(archive, worksheet_part)
                if table:
                    table_name, cell_range = table
                    return ExcelSource(
                        source_file=str(path),
                        source_sheet=sheet_name,
                        source_type="table",
                        table_name=table_name,
                        cell_range=cell_range,
                        available_sheets=tuple(available_sheets),
                    )
        return None

    def _find_table_all_in_sheet(
        self,
        archive: zipfile.ZipFile,
        worksheet_part: str,
    ) -> tuple[str, str] | None:
        rels_part = _worksheet_rels_part(worksheet_part)
        try:
            worksheet_xml = ElementTree.fromstring(archive.read(worksheet_part))
            rels_xml = ElementTree.fromstring(archive.read(rels_part))
        except KeyError:
            return None

        table_rel_ids = [
            node.attrib.get("{http://schemas.openxmlformats.org/officeDocument/2006/relationships}id")
            for node in worksheet_xml.findall(".//{http://schemas.openxmlformats.org/spreadsheetml/2006/main}tablePart")
        ]
        rels = {
            rel.attrib["Id"]: _resolve_part(worksheet_part, rel.attrib["Target"])
            for rel in rels_xml
        }
        for relation_id in table_rel_ids:
            table_part = rels.get(relation_id or "")
            if not table_part:
                continue
            table_xml = ElementTree.fromstring(archive.read(table_part))
            table_name = table_xml.attrib.get("displayName") or table_xml.attrib.get("name")
            cell_range = table_xml.attrib.get("ref")
            if table_name and table_name.lower() == "all" and cell_range:
                return table_name, cell_range
        return None

    def _to_error_records(
        self,
        batch_id: int,
        errors: list[ValidationError],
        raw_payload: dict[str, Any],
    ) -> list[ImportErrorRecord]:
        raw_payload_json = _json_dump(raw_payload) if raw_payload else None
        return [
            ImportErrorRecord(
                import_batch_id=batch_id,
                source_row_number=error.row_number,
                field_name=error.field_name,
                error_message=error.message,
                raw_value=error.raw_value,
                raw_payload_json=raw_payload_json,
            )
            for error in errors
        ]

    @staticmethod
    def _normalize_header(header: str) -> str:
        text = str(header).strip()
        if text in COLUMN_MAPPING:
            return COLUMN_MAPPING[text]
        lower = text.lower().replace("\n", " ").strip()
        return COLUMN_MAPPING.get(lower, lower.replace(" ", "_").replace("-", "_"))

    @staticmethod
    def _first(payload: dict[str, Any], *keys: str, default: Any = None) -> Any:
        for key in keys:
            value = payload.get(key)
            if not _is_empty(value):
                return value
        return default

    @staticmethod
    def _parse_date(
        value: Any,
        row_number: int,
        field_name: str,
        errors: list[ValidationError],
    ) -> str | None:
        if _is_empty(value):
            errors.append(ValidationError(row_number, field_name, "Required date is missing"))
            return None
        if isinstance(value, datetime):
            return value.date().isoformat()
        if isinstance(value, date):
            return value.isoformat()
        text = str(value).strip()
        for fmt in ("%Y-%m-%d", "%d.%m.%Y", "%d/%m/%Y", "%m/%d/%Y"):
            try:
                return datetime.strptime(text, fmt).date().isoformat()
            except ValueError:
                continue
        errors.append(ValidationError(row_number, field_name, "Invalid date format", value))
        return None

    @staticmethod
    def _parse_float(
        value: Any,
        row_number: int,
        field_name: str,
        errors: list[ValidationError],
        allow_empty: bool = False,
    ) -> float | None:
        if _is_empty(value):
            if allow_empty:
                return 0.0
            errors.append(ValidationError(row_number, field_name, "Required number is missing"))
            return None
        try:
            return float(str(value).replace(" ", "").replace(",", "."))
        except ValueError:
            errors.append(ValidationError(row_number, field_name, "Invalid number", value))
            return None

    @classmethod
    def _parse_optional_date(
        cls,
        value: Any,
        row_number: int,
        field_name: str,
        errors: list[ValidationError],
    ) -> str | None:
        if _is_empty(value):
            return None
        local_errors: list[ValidationError] = []
        parsed = cls._parse_date(value, row_number, field_name, local_errors)
        errors.extend(local_errors)
        return parsed

    @classmethod
    def _parse_optional_float(
        cls,
        value: Any,
        row_number: int,
        field_name: str,
        errors: list[ValidationError],
    ) -> float | None:
        if _is_empty(value):
            return None
        return cls._parse_float(value, row_number, field_name, errors)

    @classmethod
    def _parse_optional_positive_float(
        cls,
        value: Any,
        row_number: int,
        field_name: str,
        errors: list[ValidationError],
    ) -> float | None:
        """Parse a numeric value and normalize it to positive when Excel sends signed exports."""
        parsed = cls._parse_optional_float(value, row_number, field_name, errors)
        return abs(parsed) if parsed is not None else None

    @staticmethod
    def _parse_bool(
        value: Any,
        row_number: int | None = None,
        field_name: str = "value",
        errors: list[ValidationError] | None = None,
    ) -> bool:
        if isinstance(value, bool):
            return value
        text = str(value).strip().casefold()
        if text == "\u0434\u0430":
            return True
        if text == "\u043d\u0435\u0442":
            return False
        if text in {"1", "true", "yes", "y"}:
            return True
        if text in {"0", "false", "no", "n"}:
            return False
        if errors is not None and row_number is not None:
            errors.append(
                ValidationError(
                    row_number,
                    field_name,
                    "Boolean value must be '\u0434\u0430' or '\u043d\u0435\u0442'",
                    value,
                )
            )
        return False

    @staticmethod
    def _optional_bool(
        value: Any,
        row_number: int | None = None,
        field_name: str = "value",
        errors: list[ValidationError] | None = None,
    ) -> bool:
        if _is_empty(value):
            return False
        return ImportExcelService._parse_bool(value, row_number, field_name, errors)

    @staticmethod
    def _optional_text(value: Any) -> str | None:
        if _is_empty(value):
            return None
        return str(value).strip()

    @staticmethod
    def _optional_upper(value: Any) -> str | None:
        text = ImportExcelService._optional_text(value)
        return text.upper() if text else None

    @staticmethod
    def _is_supported_currency_code(value: str) -> bool:
        code = value.strip().upper()
        return len(code) == 3 or code in SUPPORTED_EXTENDED_CURRENCIES

    @staticmethod
    def _is_empty_row(payload: dict[str, Any]) -> bool:
        return all(_is_empty(value) for value in payload.values())

    @staticmethod
    def _build_comment(payload: dict[str, Any]) -> str | None:
        external_id = payload.get("external_deal_id")
        if _is_empty(external_id):
            return None
        return f"External deal ID: {external_id}"


def _resolve_part(base_part: str, target: str) -> str:
    if target.startswith("/"):
        return target.lstrip("/")
    return posixpath.normpath(posixpath.join(posixpath.dirname(base_part), target))


def _worksheet_rels_part(worksheet_part: str) -> str:
    directory, filename = posixpath.split(worksheet_part)
    return posixpath.join(directory, "_rels", f"{filename}.rels")


def _is_empty(value: Any) -> bool:
    if value is None:
        return True
    return str(value).strip() == ""


def _json_dump(payload: dict[str, Any]) -> str:
    return json.dumps(payload, ensure_ascii=False, default=str)
